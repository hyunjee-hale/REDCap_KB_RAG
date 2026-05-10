"""
kb_to_servicenow.py
Converts REDCap KB markdown articles to a ServiceNow-ready Excel import file.

Reads from the "kb (YAML)" folder — articles use YAML frontmatter for metadata
and standard markdown for body content.

Sheet 1 – Articles      : one row per article, fields mapped to kb_knowledge columns
Sheet 2 – Relationships : one row per cross-reference (Prerequisites + Related Topics)
"""

import re
import yaml
import markdown
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Instance configuration ────────────────────────────────────────────────────
# Paths are loaded from config.local.yaml in the repo root.
# Copy config.example.yaml → config.local.yaml and fill in your values.
# Falls back to sensible defaults relative to this script's location if the
# config file is not present (useful for running from within the repo directly).

_REPO_ROOT = Path(__file__).resolve().parent.parent
_CONFIG_PATH = _REPO_ROOT / "config.local.yaml"

def _load_config() -> dict:
    if _CONFIG_PATH.exists():
        with _CONFIG_PATH.open(encoding="utf-8") as f:
            return yaml.safe_load(f) or {}
    return {}

_cfg = _load_config()
_etl = _cfg.get("etl", {})

KB_DIR  = Path(_etl.get("kb_dir",  _REPO_ROOT / "kb (YAML)"))
OUT_FILE = Path(_etl.get("out_file", _REPO_ROOT / "ServiceNow ETL" / "REDCap_KB_ServiceNow_Import.xlsx"))

# Regex to extract RC-xxx IDs from strings like "RC-FD-02 — Online Designer"
REF_RE = re.compile(r"(RC-[A-Z0-9-]+)")


def parse_article(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")

    # --- Split YAML frontmatter from markdown body ---
    front = {}
    body_text = text.strip()
    if text.startswith("---"):
        parts = text.split("---", 2)
        if len(parts) >= 3:
            front = yaml.safe_load(parts[1]) or {}
            body_text = parts[2].strip()

    # --- Core fields from frontmatter ---
    source_id  = str(front.get("id", path.stem.split("_")[0]))
    title      = str(front.get("title", ""))
    domain     = str(front.get("domain", ""))
    version    = str(front.get("version", ""))
    last_updated = str(front.get("last_updated", ""))

    # applies_to: list → semicolon-separated string
    applies_to_raw = front.get("applies_to", [])
    if isinstance(applies_to_raw, list):
        applies_to = "; ".join(str(x) for x in applies_to_raw)
    else:
        applies_to = str(applies_to_raw)

    # --- Prerequisites ---
    prereqs_raw = front.get("prerequisites", [])
    if not isinstance(prereqs_raw, list):
        prereqs_raw = [prereqs_raw] if prereqs_raw else []
    # Filter out bare "None" entries
    prereqs_filtered = [p for p in prereqs_raw if str(p).strip().lower() != "none"]
    prerequisite_raw = "; ".join(str(p) for p in prereqs_raw)
    prereq_ids = []
    for p in prereqs_filtered:
        prereq_ids.extend(REF_RE.findall(str(p)))

    # --- Related topics ---
    related_raw = front.get("related", [])
    if not isinstance(related_raw, list):
        related_raw = [related_raw] if related_raw else []
    related_parts = []
    related_ids  = []
    for r in related_raw:
        if isinstance(r, dict):
            rid    = r.get("id", "")
            rtitle = r.get("title", "")
            related_parts.append(f"{rid} — {rtitle}" if rid and rtitle else rid or rtitle)
            if rid:
                related_ids.append(rid)
        else:
            s = str(r)
            related_parts.append(s)
            related_ids.extend(REF_RE.findall(s))
    related_topics_raw = "; ".join(related_parts)

    # --- Convert markdown body to HTML ---
    html_body = markdown.markdown(
        body_text,
        extensions=["tables", "fenced_code", "nl2br"]
    )

    return {
        "source_id":          source_id,
        "short_description":  title,
        "category":           domain,
        "applies_to":         applies_to,
        "version":            version,
        "last_updated":       last_updated,
        "author":             "",          # not present in YAML frontmatter
        "prerequisite_raw":   prerequisite_raw,
        "related_topics_raw": related_topics_raw,
        "workflow_state":     "draft",
        "kb_knowledge_base":  "",          # to be filled by ServiceNow admin
        "text":               html_body,
        "_prereq_ids":        prereq_ids,
        "_related_ids":       related_ids,
        "_path":              str(path.name),
    }


def style_header_row(ws, row=1, color="1F4E79"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_border(ws):
    thin = Side(style="thin", color="CCCCCC")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_excel(articles: list[dict]):
    wb = Workbook()

    # ── Sheet 1: Articles ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Articles"

    art_headers = [
        "source_id",          # your RC-xxx ID — store in u_source_id custom field
        "short_description",  # article title → kb_knowledge.short_description
        "category",           # domain → kb_knowledge.category
        "kb_knowledge_base",  # leave blank; fill in after admin configures SN
        "workflow_state",     # draft / published
        "author",
        "version",
        "last_updated",
        "applies_to",         # maps to kb_knowledge.meta_description or custom field
        "prerequisite_raw",   # human-readable; for reference
        "related_topics_raw", # human-readable; for reference
        "text",               # HTML body → kb_knowledge.text
    ]

    ws1.append(art_headers)
    style_header_row(ws1, row=1, color="1F4E79")
    ws1.row_dimensions[1].height = 22

    for art in articles:
        ws1.append([art.get(h, "") for h in art_headers])

    # Column widths
    widths = {
        "A": 14, "B": 45, "C": 20, "D": 25, "E": 14,
        "F": 20, "G": 10, "H": 14, "I": 40, "J": 50, "K": 50, "L": 80
    }
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w

    # Wrap text and top-align data rows
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column != 12))
            cell.font = Font(name="Arial", size=10)
        ws1.row_dimensions[cell.row].height = 40

    # Freeze header
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Relationships ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Relationships")

    rel_headers = [
        "source_id",           # article that contains the reference
        "source_title",
        "relationship_type",   # Prerequisite | Related Topic
        "target_id",           # RC-xxx being referenced
        "target_title",        # resolved if that article is also in the set
        "notes",
    ]

    ws2.append(rel_headers)
    style_header_row(ws2, row=1, color="375623")
    ws2.row_dimensions[1].height = 22

    # Build a lookup: source_id → title
    title_map = {a["source_id"]: a["short_description"] for a in articles}

    for art in articles:
        sid = art["source_id"]
        stitle = art["short_description"]
        for tid in art["_prereq_ids"]:
            ws2.append([
                sid, stitle, "Prerequisite", tid,
                title_map.get(tid, "⚠ Not in current export"),
                "Must be read before this article"
            ])
        for tid in art["_related_ids"]:
            ws2.append([
                sid, stitle, "Related Topic", tid,
                title_map.get(tid, "⚠ Not in current export"),
                ""
            ])

    rel_widths = {"A": 16, "B": 45, "C": 18, "D": 16, "E": 45, "F": 35}
    for col, w in rel_widths.items():
        ws2.column_dimensions[col].width = w

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=10)
        ws2.row_dimensions[cell.row].height = 30

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Instructions ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Import Instructions")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 80

    instructions = [
        ("REDCap KB → ServiceNow Import Guide", ""),
        ("", ""),
        ("SHEET: Articles", ""),
        ("source_id", "Store in a custom field (u_source_id) so articles remain findable by RC-xxx ID after import."),
        ("short_description", "Maps to kb_knowledge.short_description (the article title)."),
        ("category", "Maps to kb_knowledge.category. Your SN admin may need to create matching category records first."),
        ("kb_knowledge_base", "Fill in the sys_id or name of your target Knowledge Base before importing."),
        ("workflow_state", "Use 'draft' initially; publish after review. Accepted values: draft, review, published, retired."),
        ("author", "Must match a valid ServiceNow user (by name or sys_id). Update before import."),
        ("text", "HTML body. Paste into kb_knowledge.text. Verify formatting in SN preview after import."),
        ("", ""),
        ("SHEET: Relationships", ""),
        ("How to use", "After Pass 1 import, use this sheet to wire up Related Articles links in ServiceNow."),
        ("⚠ Not in current export", "Target article was referenced but its source .md file was not in this export batch. Import it separately."),
        ("Two-pass approach", "1) Import Articles sheet → note the KB article numbers SN assigns.\n2) Use Relationships sheet + those numbers to create related-article links via SN admin or script."),
        ("", ""),
        ("RECOMMENDED CUSTOM FIELD", ""),
        ("u_source_id", "Add this custom field to kb_knowledge in ServiceNow. Store the RC-xxx ID here. Enables reliable lookups and future re-imports without duplication."),
    ]

    ws3.append(["Field / Topic", "Guidance"])
    style_header_row(ws3, row=1, color="4A4A4A")
    ws3.row_dimensions[1].height = 22

    for i, (field, guidance) in enumerate(instructions, start=2):
        ws3.cell(row=i, column=1, value=field)
        ws3.cell(row=i, column=2, value=guidance)
        if field in ("SHEET: Articles", "SHEET: Relationships", "RECOMMENDED CUSTOM FIELD"):
            for col in [1, 2]:
                ws3.cell(row=i, column=col).font = Font(name="Arial", size=10, bold=True)
                ws3.cell(row=i, column=col).fill = PatternFill("solid", start_color="D9E1F2")
        else:
            for col in [1, 2]:
                ws3.cell(row=i, column=col).font = Font(name="Arial", size=10)
        ws3.row_dimensions[i].height = 36

    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(OUT_FILE)
    print(f"Saved: {OUT_FILE}")
    return OUT_FILE


if __name__ == "__main__":
    md_files = sorted(KB_DIR.glob("RC-*.md"))
    print(f"Found {len(md_files)} KB articles:")
    for f in md_files:
        print(f"  {f.name}")

    articles = [parse_article(f) for f in md_files]

    print("\nArticle summary:")
    for a in articles:
        print(f"  {a['source_id']:12s} | {a['category']:20s} | prereqs={a['_prereq_ids']} | related={a['_related_ids']}")

    build_excel(articles)
    print("Done.")
